import os
import django

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.sys.path.append(BASE_DIR)
os.environ['DJANGO_SETTINGS_MODULE'] = 'core.settings'
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

import openpyxl
from main.models import OT, SIFAT, UmumiyTurkum

path = "E:\MorphApp\KompLingvistika\main\dataset.xlsx"
UmumiyTurkum.objects.filter().delete()
wb_obj = openpyxl.load_workbook(path,data_only=True)
sheet_obj = wb_obj.active

print("Boshlandi")
for i in range(2,1000):
    ot = sheet_obj.cell(row=i, column=1).value
    sifat = sheet_obj.cell(row=i, column=2).value
    son = sheet_obj.cell(row=i, column=3).value
    fel = sheet_obj.cell(row=i, column=4).value
    ravish = sheet_obj.cell(row=i, column=5).value
    olmosh = sheet_obj.cell(row=i, column=6).value

    if ot:
        UmumiyTurkum.objects.create(
            word = ot, type_is = '1'
        )
    if sifat:
        UmumiyTurkum.objects.create(
            word = sifat, type_is = '2'
        )
    if son:
        UmumiyTurkum.objects.create(
            word = son, type_is = '3'
        )
    if fel:
        UmumiyTurkum.objects.create(
            word = fel, type_is = '4'
        )
    if ravish:
        UmumiyTurkum.objects.create(
            word = ravish.split('.')[1].lstrip().rstrip(), type_is = '5'
        )
    if olmosh:
        UmumiyTurkum.objects.create(
            word = olmosh.split('.')[1].lstrip().rstrip(), type_is = '6'
        )

print("Tugadi")